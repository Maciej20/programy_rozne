# Program odtwarza po kolei filmy z folderu "mov",
# a z pliku "cfg.txt" pobiera dane o czterch punktach i rysuje kreski na ekranie.
# Kreski rozmieszczone są w znanej odległośći rzeczywistej (na podstawie oznakowania poziomego na jezdni)
# Pracownicy klikają odpowiednie przyciski na klawiaturze kiedy pojazd przekroczy linie
# Liczbę klatek na sekundę odczytywana jest również z pliku "cfg.txt" i dzięki temu można wyliczyć rzeczywisty czas 
# Wyniki przechowywane są na listach, a po zakończeniu filmu zapisywane są do pliku "data.xlsx"
# W pliku "data.xlsx" znajduje się także arkusz "log", który zbiera dane o tym, ile pracownik poświęcił czasu na każdy film

from datetime import datetime, time
import os
import cv2
from openpyxl.descriptors.base import String
import pygame
from openpyxl import load_workbook
from pygame.locals import ( 
    KEYDOWN,
    K_j,
    K_k,
    K_u,
    K_i,
    K_q,
    K_w,
    K_a,
    K_s,
    K_z,
    K_x,
    K_ESCAPE,
)


DODAJ_LEKKIE_1 = "j"
DODAJ_LEKKIE_2 = "k"
DODAJ_CIĘŻKIE_1 = "J"
DODAJ_CIĘŻKIE_2 = "K"
filmy = []

directory = os.path.realpath(__file__).rstrip("SpeedometerV2.py") + "mov\\"
for filename in os.listdir(directory):
    if filename.endswith(".avi"):
        filmy.append(filename)

cfg = []

with open('cfg.txt') as f:
    line = f.readline()
    while line:
        try:
            line = int(f.readline()[2:].rstrip())
            cfg.append(line)
        except ValueError:
            break


k1_l, k2_l, k1_c, k2_c = 0, 0, 0, 0
frame_counter = 0
FPS = cfg[-1]
dzialaj = True
licznik = 0
pygame.init()

def put_duration_to_excel(starttime: datetime, pausetime: float, action: String):
    wb = load_workbook("data.xlsx") 
    ws = wb["log"]
    x = 2
    while True:
        if ws.cell(row = x, column= 1).value is not None:
            x += 1
        else:
            break
    
    ws.cell(row = x, column = 1, value = starttime)
    time_now = datetime.now()
    duration = time_now - starttime
    duration_in_seconds = 0
    try:
        duration_in_seconds += duration.hours * 3600
    except AttributeError:
        pass
    try:
        duration_in_seconds += duration.minutes* 60
    except AttributeError:
        pass
    try:
        duration_in_seconds += duration.seconds
    except AttributeError:
        pass
    duration_in_seconds += duration.microseconds/1000000

    ws.cell(row = x, column = 2, value = round(duration_in_seconds, 2))
    ws.cell(row = x, column = 3, value = round(pausetime, 2))
    ws.cell(row = x, column = 4, value = action)
    wb.save("data.xlsx")
    

def put_list_to_excel(data: list, type: String, quarter: String):
    hour = int(quarter[:2])
    if type == "l":
        lekkie = True
    else:
        lekkie = False
    wb = load_workbook("data.xlsx")
    ws = wb["arkusz"]

    if lekkie:
        x = 3
        while True:
            if ws.cell(row = x, column= 2*hour+1).value is not None:
                x += 1
            else:
                break
        for index, val in enumerate(data):
            ws.cell(row = index + x, column = 2*hour+1, value = val)
    else:
        y = 3
        while True:
            if ws.cell(row = y, column= 2*hour+2).value is not None:
                y += 1
            else:
                break
        for index, val in enumerate(data):
            ws.cell(row = index + y, column = 2*hour+2, value = val)
    wb.save("data.xlsx")

while licznik < len(filmy) and dzialaj == True:
    frame_counter = 0
    nr = licznik
    cap = cv2.VideoCapture(f"{directory}{filmy[nr]}")
    success, img = cap.read()
    shape = [1600,720]
    wn = pygame.display.set_mode(shape, pygame.RESIZABLE)
    clock = pygame.time.Clock()
    font = pygame.font.Font('freesansbold.ttf', 32)
    text = font.render(str(0), True, (0,0,0), (0,255,0))
    counter = 0
    tick = 20
    lekkie = []
    ciezkie = []
    pauza = False

    kwadrans = filmy[nr][-9:-4]
    starttime = datetime.now()
    pauza_timestamps = []
    start_timestamps = []

    while success:
        clock.tick(tick)
        success, img = cap.read()
        cv2.line(img, (cfg[0], cfg[1]), (cfg[2], cfg[3]), (0, 255, 0), thickness=3)
        cv2.line(img, (cfg[4], cfg[5]), (cfg[6], cfg[7]), (0, 0, 255), thickness=3)

        while pauza:
            for event in pygame.event.get():
                if event.type == KEYDOWN and event.key == K_s:
                    pauza = False
                    start_timestamps.append(datetime.now())
                if event.type == KEYDOWN and event.key == K_ESCAPE:
                    pauza = False
                    success = False
                if event.type == pygame.QUIT:
                    pauza = False
                    success = False
                    dzialaj = False
                    break

        tempo = tick/5
        frame_counter +=1
        text = font.render(str(tempo), True, (255,255,255), (0,0,0))
        text2 = font.render(str(len(lekkie)), True, (255,255,255), (0,0,0))
        text3 = font.render(str(len(ciezkie)), True, (255,255,255), (0,0,0))
        text4 = font.render(str(kwadrans), True, (255,255,255), (0,0,0))
        text5 = font.render(str(frame_counter), True, (255,255,255), (0,0,0))
        wn.blit(text, (1300,150))
        wn.blit(text2, (1300,200))
        wn.blit(text3, (1300,250))
        wn.blit(text4, (1300,50))
        wn.blit(text5, (1300,100))

        for event in pygame.event.get():
            if event.type == KEYDOWN and event.key == K_j:
                k1_l = frame_counter
            if event.type == KEYDOWN and event.key == K_k:
                k2_l = frame_counter
                lekkie.append((k2_l - k1_l)/FPS)
            if event.type == KEYDOWN and event.key == K_u:
                k1_c = frame_counter
            if event.type == KEYDOWN and event.key == K_i:
                k2_c = frame_counter
                ciezkie.append((k2_c - k1_c)/FPS)

            if event.type == KEYDOWN and event.key == K_q:
                tick +=1
            if event.type == KEYDOWN and event.key == K_w:
                tick -= 1
            if event.type == KEYDOWN and event.key == K_a:
                pauza = True
                pauza_timestamps.append(datetime.now())

            if event.type == KEYDOWN and event.key == K_ESCAPE:
                success = False
            if event.type == KEYDOWN and event.key == K_z:
                licznik -= 1
                success = False
            if event.type == KEYDOWN and event.key == K_x:
                licznik -= 2
                success = False

            if event.type == pygame.VIDEORESIZE:
                # There's some code to add back window content here.
                wn = pygame.display.set_mode((event.w, event.h), pygame.RESIZABLE)
            if event.type == pygame.QUIT:
                dzialaj = False
                success = False
                break
            
        try:
            wn.blit(pygame.image.frombuffer(img.tobytes(), (1280,720), "BGR"), (0, 0))
            pygame.display.update()
        except:
            pass


    pausetimes = list(map(lambda x, y: x-y, start_timestamps, pauza_timestamps))
    pausetime = 0
    for x in pausetimes:
        try:
            pausetime += x.minutes * 60
        except AttributeError:
            pass
        try:
            pausetime += x.hours * 3600
        except AttributeError:
            pass
        try:
            pausetime += x.seconds
        except AttributeError:
            pass
        pausetime += x.microseconds/1000000

    put_list_to_excel(lekkie, "l", kwadrans)
    put_list_to_excel(ciezkie, "c", kwadrans)
    put_duration_to_excel(starttime, pausetime, kwadrans)

    licznik += 1

pygame.quit()