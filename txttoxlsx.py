#Program przetwarza plik txt będący wynikiem pracy pracowników oglądających filmy 
#na czytelną dla człowieka i przystępniejszą w późniejszej obróbce danych tabelkę w excelu

from openpyxl import load_workbook
from collections import Counter

while True:
    filename = input("Podaj nazwę pliku: ")
    if filename == "q":
        break
    groupname = input("Podaj grupę występującą w pliku: ")
    lines = []
    with open(f"{filename}.txt", encoding = 'utf-8') as txt:
        for line in txt:
            lines.append(line.rstrip())
    
    wb = load_workbook("clear.xlsx")
    ws = wb["template"]

    x = 0
    while x < 20:
        if lines[x] == "0000":
            x += 1
            break
        else:
            x += 1
    
    for a in range (x, len(lines)):
        if lines[a].isdigit() or lines[a].startswith("g") == False:
            now_row = (a - x)/2 + 3
            c = Counter(lines[a])
            if groupname == "j":
                ws.cell(row = now_row, column = 2, value = c["j"])
                ws.cell(row = now_row, column = 3, value = c["k"])
                ws.cell(row = now_row, column = 4, value = c["l"])
                ws.cell(row = now_row, column = 5, value = c[";"])
                ws.cell(row = now_row, column = 6, value = c["J"])
                ws.cell(row = now_row, column = 7, value = c["K"])
                ws.cell(row = now_row, column = 8, value = c["L"])
                ws.cell(row = now_row, column = 9, value = c[":"])
            if groupname == "f":
                ws.cell(row = now_row, column = 2, value = c["f"])
                ws.cell(row = now_row, column = 3, value = c["d"])
                ws.cell(row = now_row, column = 4, value = c["s"])
                ws.cell(row = now_row, column = 5, value = c["a"])
                ws.cell(row = now_row, column = 6, value = c["F"])
                ws.cell(row = now_row, column = 7, value = c["D"])
                ws.cell(row = now_row, column = 8, value = c["S"])
                ws.cell(row = now_row, column = 9, value = c["A"])

    ws.cell(row = 3, column = 1, value = 0)
    wb.save(f"wyniki{filename.lstrip('plik_pomiarowy')}.xlsx")